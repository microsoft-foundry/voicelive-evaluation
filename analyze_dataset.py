import os, sys
try:
    import openpyxl
except ImportError:
    os.system(sys.executable + " -m pip install openpyxl --quiet")
    import openpyxl

base = r"C:\Localrepos\voicelive-evaluation\evaluation_harness\local_datasets\BingChatDataSetWGroundTruth"

# Collect all xlsx files
all_files = []
for root, dirs, files in os.walk(base):
    for f in files:
        if f.endswith('.xlsx'):
            all_files.append(os.path.join(root, f))

print(f"Total files found: {len(all_files)}\n")

# Analyze each file
summary = []
all_comments = {}
time_sensitive_with_gt = []
time_sensitive_without_gt = []
unanswerable_with_gt = []
no_judge_selected = []
gt_contradicts_all = []

for fpath in sorted(all_files):
    rel = os.path.relpath(fpath, base)
    parts = rel.split(os.sep)
    delivery = parts[0]
    lang = parts[1] if len(parts) > 1 else "?"
    fname = parts[-1]
    
    try:
        wb = openpyxl.load_workbook(fpath, read_only=True)
    except Exception as e:
        print(f"WARNING: Skipping {rel} - {e}")
        continue
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        
        headers = [str(h).strip() if h else '' for h in rows[0]]
        
        # Find key columns by name
        def find_col(keywords):
            for i, h in enumerate(headers):
                hl = h.lower()
                if all(k in hl for k in keywords):
                    return i
            return None
        
        gt_col = find_col(['human', 'final', 'gt']) or find_col(['human_final_gt']) or find_col(['groundtruth']) or find_col(['ground_truth'])
        comment_col = find_col(['comment'])
        judge_col = find_col(['judge', 'selected'])
        judge_reason_col = find_col(['judge', 'reason'])
        transcription_col = find_col(['transcription'])
        
        # LLM output columns
        llm_cols = [i for i, h in enumerate(headers) if h.lower().startswith('output_')]
        
        data_rows = rows[1:]
        total = len(data_rows)
        
        # Count empty GT
        empty_gt = 0
        non_empty_gt = 0
        
        # Comment analysis
        comment_counts = {}
        
        for row in data_rows:
            # GT analysis
            gt_val = row[gt_col] if gt_col is not None and gt_col < len(row) else None
            gt_empty = (gt_val is None or str(gt_val).strip() == '')
            if gt_empty:
                empty_gt += 1
            else:
                non_empty_gt += 1
            
            # Comment analysis
            comment_val = row[comment_col] if comment_col is not None and comment_col < len(row) else None
            if comment_val and str(comment_val).strip():
                c = str(comment_val).strip()
                comment_counts[c] = comment_counts.get(c, 0) + 1
                
                c_lower = c.lower()
                
                # Time-sensitive checks
                if 'time' in c_lower or 'sensitive' in c_lower or 'dated' in c_lower or 'temporal' in c_lower:
                    if gt_empty:
                        time_sensitive_without_gt.append((lang, fname, str(row[transcription_col] if transcription_col is not None else '')[:80]))
                    else:
                        time_sensitive_with_gt.append((lang, fname, str(row[transcription_col] if transcription_col is not None else '')[:80], str(gt_val)[:80]))
                
                # Unanswerable with GT check
                if ('unanswer' in c_lower or 'no answer' in c_lower or 'cannot' in c_lower) and not gt_empty:
                    unanswerable_with_gt.append((lang, fname, str(row[transcription_col] if transcription_col is not None else '')[:80], str(gt_val)[:80]))
            
            # Judge selection analysis
            judge_val = row[judge_col] if judge_col is not None and judge_col < len(row) else None
            if judge_val is None or str(judge_val).strip() == '':
                if not gt_empty:
                    no_judge_selected.append((lang, fname))
            
            # GT contradicts all LLMs check (only if GT is non-empty and all LLM outputs exist)
            if not gt_empty and llm_cols:
                gt_str = str(gt_val).strip().lower()
                all_llm_empty = all(
                    (row[c] is None or str(row[c]).strip() == '') 
                    for c in llm_cols if c < len(row)
                )
                # Don't flag if all LLMs are empty
                if not all_llm_empty and len(gt_str) > 10:
                    # Check if GT is very different from all LLM outputs
                    # Simple heuristic: check if GT shares less than 20% words with any LLM output
                    gt_words = set(gt_str.split())
                    if len(gt_words) >= 3:
                        max_overlap = 0
                        for c in llm_cols:
                            if c < len(row) and row[c]:
                                llm_words = set(str(row[c]).strip().lower().split())
                                if llm_words:
                                    overlap = len(gt_words & llm_words) / len(gt_words)
                                    max_overlap = max(max_overlap, overlap)
                        if max_overlap < 0.15:
                            gt_contradicts_all.append((lang, fname, str(row[transcription_col] if transcription_col is not None else '')[:60]))
        
        # Store unique comments
        for c, cnt in comment_counts.items():
            key = c[:100]
            if key not in all_comments:
                all_comments[key] = {'count': 0, 'langs': set()}
            all_comments[key]['count'] += cnt
            all_comments[key]['langs'].add(lang)
        
        summary.append({
            'delivery': delivery[:20],
            'lang': lang,
            'file': fname[:50],
            'total': total,
            'empty_gt': empty_gt,
            'non_empty_gt': non_empty_gt,
            'pct_empty': f"{100*empty_gt/total:.1f}%" if total > 0 else "N/A",
            'has_comments': sum(comment_counts.values()),
        })
    
    wb.close()

# Print summary table
print("=" * 120)
print("FILE-BY-FILE SUMMARY")
print("=" * 120)
print(f"{'Delivery':<22} {'Lang':<8} {'File':<52} {'Total':>6} {'EmptyGT':>8} {'%Empty':>7} {'Comments':>8}")
print("-" * 120)
for s in summary:
    print(f"{s['delivery']:<22} {s['lang']:<8} {s['file']:<52} {s['total']:>6} {s['empty_gt']:>8} {s['pct_empty']:>7} {s['has_comments']:>8}")

# Print unique comments
print(f"\n{'=' * 80}")
print("UNIQUE COMMENT VALUES (across all files)")
print(f"{'=' * 80}")
for c, info in sorted(all_comments.items(), key=lambda x: -x[1]['count']):
    print(f"  [{info['count']:>4}x] [{', '.join(sorted(info['langs']))}] {c}")

# Print inconsistency findings
print(f"\n{'=' * 80}")
print("TIME-SENSITIVE QUESTIONS WITH GT (potential inconsistency)")
print(f"{'=' * 80}")
if time_sensitive_with_gt:
    for lang, fn, q, gt in time_sensitive_with_gt[:15]:
        print(f"  [{lang}] Q: {q}")
        print(f"          GT: {gt}")
else:
    print("  None found")

print(f"\nTime-sensitive WITHOUT GT: {len(time_sensitive_without_gt)}")
print(f"Time-sensitive WITH GT: {len(time_sensitive_with_gt)}")

print(f"\n{'=' * 80}")
print("UNANSWERABLE/NO-ANSWER TAGGED BUT HAS GT (inconsistency)")
print(f"{'=' * 80}")
if unanswerable_with_gt:
    for lang, fn, q, gt in unanswerable_with_gt[:15]:
        print(f"  [{lang}] Q: {q}")
        print(f"          GT: {gt}")
else:
    print("  None found")

print(f"\n{'=' * 80}")
print(f"GT POTENTIALLY CONTRADICTS ALL LLM OUTPUTS: {len(gt_contradicts_all)} cases")
print(f"{'=' * 80}")
if gt_contradicts_all:
    for lang, fn, q in gt_contradicts_all[:10]:
        print(f"  [{lang}] {fn[:30]} | {q}")

print(f"\nNo Judge Selected but GT exists: {len(no_judge_selected)} cases")
